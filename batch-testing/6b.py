from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
wb = load_workbook("test_data.xlsx")
sheet = wb.active
driver = webdriver.Chrome()
wait = WebDriverWait(driver, 10)
for row in sheet.iter_rows(min_row=2, values_only=True):
    username_data, password_data = row
    driver.get("https://the-internet.herokuapp.com/login")
    username = wait.until(EC.visibility_of_element_located((By.ID, "username")))
    password = wait.until(EC.visibility_of_element_located((By.ID, "password")))
    username.clear()
    password.clear()
    username.send_keys(username_data)
    password.send_keys(password_data)
    driver.find_element(By.CSS_SELECTOR, "button[type='submit']").click()
    try:
        success_msg = wait.until(EC.presence_of_element_located((By.ID, "flash")))
        print(f"{username_data} → RESULT: {success_msg.text.strip()}")
    except:
        print(f"{username_data} → RESULT: Not found")
driver.quit()
