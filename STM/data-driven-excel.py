from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
import time
print("Data Driven Test - Excel")
driver = webdriver.Chrome()
wb = load_workbook("test_data.xlsx")
sheet = wb.active
try:
    for i, row in enumerate(sheet.iter_rows(min_row=1, max_col=1), start=1):
        data = row[0].value
        if data is None:
            continue
        print(f"\nTest Case {i}")
        print("Input:", data)
        driver.get("https://www.google.com")
        try:
            consent = WebDriverWait(driver, 5).until(
                EC.element_to_be_clickable((By.XPATH, "//button[contains(., 'Accept')]"))
            )
            consent.click()
        except:
            pass
        search = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.NAME, "q"))
        )
        search.clear()
        search.send_keys(str(data))
        search.send_keys(Keys.RETURN)
        print("Search completed")
        time.sleep(2)

finally:
    driver.quit()
