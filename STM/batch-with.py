from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
driver = webdriver.Chrome()
wait = WebDriverWait(driver, 10)
wb = load_workbook("test_data.xlsx")
sheet = wb.active
for row in sheet.iter_rows(min_row=2, values_only=True):
    username_data, password_data = row[:2]
    driver.get("https://the-internet.herokuapp.com/login")
    wait.until(EC.element_to_be_clickable((By.ID, "username"))).send_keys(username_data)
    wait.until(EC.element_to_be_clickable((By.ID, "password"))).send_keys(password_data)
    driver.find_element(By.CSS_SELECTOR, "button[type='submit']").click()
    wait.until(lambda d: "secure" in d.current_url or "login" in d.current_url)
    msg = wait.until(EC.visibility_of_element_located((By.ID, "flash"))).text
    print(username_data, "→", msg.strip())
driver.quit()
