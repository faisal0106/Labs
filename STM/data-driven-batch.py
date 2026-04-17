from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait as W
from selenium.webdriver.support import expected_conditions as E
from openpyxl import load_workbook
import os
f = os.path.join(os.path.dirname(__file__), "test_data.xlsx")
s = load_workbook(f).active
d = webdriver.Chrome()
w = W(d,10)
for i,r in enumerate(s.iter_rows(min_row=2),2):
    d.get("https://the-internet.herokuapp.com/login")
    w.until(E.visibility_of_element_located((By.ID,"username"))).send_keys(r[0].value)
    w.until(E.visibility_of_element_located((By.ID,"password"))).send_keys(r[1].value)
    d.find_element(By.CSS_SELECTOR,"button[type='submit']").click()
    try:
        res = "PASS" if "secure area" in w.until(E.visibility_of_element_located((By.ID,"flash"))).text else "FAIL"
    except:
        res = "ERROR"
    s.cell(i,3).value = res
    print(r[0].value,"→",res)
s.parent.save(f)
d.quit()
