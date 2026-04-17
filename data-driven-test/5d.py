from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from openpyxl import load_workbook
print("Data Driven Test - Excel")
driver = webdriver.Chrome()
wb = load_workbook("test_data.xlsx")
sheet = wb.active
for i, row in enumerate(sheet.iter_rows(min_row=1, max_col=1), start=1):
    data = row[0].value
    print(f"\nTest Case {i}")
    print("Input:", data)
    driver.get("https://www.google.com")
    search = WebDriverWait(driver, 10).until(
        lambda d: d.find_element(By.NAME, "q")
    )
    search.send_keys(str(data))
    search.submit()
    print("Search completed")
driver.quit()
